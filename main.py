import openpyxl #エクセル操作
import shutil #フォルダ操作
from getpass import getpass #パスワード取得

#エクセルとの紐付け
kamoku = openpyxl.load_workbook('./kamoku.xlsx')
data = kamoku['data']
data_t = kamoku['data_t']

userdata = openpyxl.load_workbook('./userdata.xlsx')
userlist = userdata['userlist']

#基本機能
class func:

    #テキストの表示
    def txt(self, name):
        with open(f'txt/{name}.txt', 'r') as tmp:
            print(tmp.read(), end='')

    #excel/userdata.xlsx内のuserlistに保存
    def add_userlist(self, data, col, row):
        userlist[f'{col}{row}'] = data
        userdata.save('./userdata.xlsx')

func = func()

#ユーザーデータに関わる機能
class user:

    #ユーザーの新規登録
    def new(self):
        i = 2
        print('希望するユーザーネームを入力してください: ', end="")
        username = input()
        password = getpass()
        #ユーザーネームの重複がないか確認、重複の場合0を返す
        while(userlist[f'b{i}'].value != None):
            if username == userlist[f'b{i}'].value:
                print('ERROR: このユーザーネームは使われています')
                return 0
            i = i + 1
        #重複がなければ保存とアカウント用のディレクトリを作成し1を返す
        print('COMPLETED: ' + username + 'さん、ユーザー登録が完了しました')
        func.add_userlist(i-1, 'a', i)
        func.add_userlist(username, 'b', i)
        func.add_userlist(password, 'c', i)
        shutil.copytree('./copy', './data/'+f'{username}')
        return 1

    #ログイン
    def login(self):
        i = 2
        print('ユーザーネームを入力してください: ', end="")
        username = input()
        password = getpass()
        while(userlist[f'b{i}'].value != None):
            if username == userlist[f'b{i}'].value and password == userlist[f'c{i}'].value:
                return [userlist[f'a{i}'].value, username, password]
            i = i + 1
        print('ERROR: ユーザーネームかパスワードが間違っています')
        return 0

    #ユーザーデータの削除
    def delete(self):
        i = 2
        print('削除したいアカウントのユーザーネームを入力してください: ', end="")
        username = input()
        password1 = getpass()
        while(userlist[f'b{i}'].value != None):
            if username == userlist[f'b{i}'].value and password1 == userlist[f'c{i}'].value:
                print('本当に削除しますか？削除する場合、【y】と入力してください: ', end="")
                c = input()
                if c == 'y':
                    print('確認のためもう一度パスワードを入力してください')
                    password2 = getpass()
                    if password1 == password2:
                        print('アカウントを削除します')
                        #アカウント削除の処理をする
                        userlist.delete_rows(i)
                        shutil.rmtree('./data/'+f'{username}')
                        userdata.save('excel/userdata.xlsx')
                        return 1
                    else:
                        print('パスワードが間違っています')
                        return 0
                else:
                    print('スタートメニューに戻ります')
                    return 2
            i = i + 1
        print('一致するデータがありませんでした、スタートメニューに戻ります')
        return 2

    #コースの入力
    def corse(self, input1, inputsave):
        while(True):
            print('コースを入力してください')
            for i in range(1,5):
                print(input1[f'd{i}'].value,': ',input1[f'e{i}'].value)
            input1['b1'].value = int(input())
            for i in range(0,4):
                if input1['b1'].value == i:
                    inputsave()
                    print('COMPLETED: 正常に登録されました')
                    return 0
            input1['b1'].value = None
            print('ERROR：もう一度最初から入力してください')

    #コースの表示
    def corse_show(self, input1):
        for i in range(1, 5):
            if input1[f'b1'].value == input1[f'd{i}'].value:
                return input1[f'e{i}'].value

    #学年の入力
    def grade(self, input1, inputsave):
        while(True):
            print('学年を入力してください')
            input1['b2'].value = int(input())
            if input1['b2'].value in {1, 2, 3, 4}:
                inputsave()
                print('COMPLETED: 正常に登録されました')
                return 0
            input1['b1'].value = None
            print('ERROR：もう一度最初から入力してください')

    #学年の表示
    def grade_show(self, input1):
        return str(input1['b2'].value)
    
    #教職の入力
    def tp(self, input1, inputsave):
        print('受けているものには1、そうでないものには0を入力してください')
        for i in range(3, 8):
            print(input1[f'a{i}'].value + ':', end="")
            while(True):
                input1[f'b{i}'].value = int(input())
                if input1[f'b{i}'].value in {0, 1}:
                    break
                print('ERROR：もう一度入力し直してください')
        inputsave()
        print('COMPLETED: 正常に登録されました')

    #教職の表示
    def tp_show(self, input1):
        c = 0
        for i in range(3, 8):
            if input1[f'b{i}'].value == 1:
                print(input1[f'a{i}'].value + " ", end="")
                c = c + 1
        if c == 0:
            print('なし')
        print('')
user = user()

class calc:
    
    #学科GPA
    def gpa(self, input2):
        sum1 = 0
        sum2 = 0
        for i in range(2,166):
            sum1 = sum1 + data[f'c{i}'].value * (input2[f'd{i}'].value*4 + input2[f'e{i}'].value*3 + input2[f'f{i}'].value*2 + input2[f'g{i}'].value*1)
            sum2 = sum2 + data[f'c{i}'].value * (input2[f'd{i}'].value + input2[f'e{i}'].value + input2[f'f{i}'].value + input2[f'g{i}'].value + input2[f'h{i}'].value)
        return float(sum1/sum2)

    #GPA
    def gpaall(self, input2, input3, input4):
        sum1 = 0
        sum2 = 0
        #input2の入力
        for i in range(2,166):
            sum1 = sum1 + data[f'c{i}'].value * (input2[f'd{i}'].value*4 + input2[f'e{i}'].value*3 + input2[f'f{i}'].value*2 + input2[f'g{i}'].value*1)
            sum2 = sum2 + data[f'c{i}'].value * (input2[f'd{i}'].value + input2[f'e{i}'].value + input2[f'f{i}'].value + input2[f'g{i}'].value + input2[f'h{i}'].value)
        #input3の入力
        for i in range(2, 23):
            sum1 = sum1 + data_t[f'c{i}'].value * (input3[f'd{i}'].value*4 + input3[f'e{i}'].value*3 + input3[f'f{i}'].value*2 + input3[f'g{i}'].value*1)
            sum2 = sum2 + data_t[f'c{i}'].value * (input3[f'd{i}'].value + input3[f'e{i}'].value + input3[f'f{i}'].value + input3[f'g{i}'].value + input3[f'h{i}'].value)
        #input4の入力
        i = 2
        while(input4[f'b{i}'].value != None):
            sum1 = sum1 + input4[f'b{i}'].value * (input4[f'd{i}'].value*4 + input4[f'e{i}'].value*3 + input4[f'f{i}'].value*2 + input4[f'g{i}'].value*1)
            sum2 = sum2 + input4[f'b{i}'].value * (input4[f'd{i}'].value + input4[f'e{i}'].value + input4[f'f{i}'].value + input4[f'g{i}'].value + input4[f'h{i}'].value + input4[f'i{i}'].value + input4[f'j{i}'].value)
            i = i + 1
        return float(sum1/sum2)

    #以下は卒業研究判定。tmp=0...履修予定を含まない / tmp=1...履修予定を含む
    #卒業研究(1),(2)
    def foundation(self, input2, tmp):
        sum1 = 0
        for i in range(2,23):
            sum1 = sum1 + data[f'j{i}'].value * data[f'c{i}'].value * (input2[f'd{i}'].value + input2[f'e{i}'].value + input2[f'f{i}'].value + input2[f'g{i}'].value + input2[f'c{i}'].value*tmp)
        sum2 = 0
        for i in range(23,35):
            sum1 = sum1 + data[f'k{i}'].value * data[f'c{i}'].value * (input2[f'd{i}'].value + input2[f'e{i}'].value + input2[f'f{i}'].value + input2[f'g{i}'].value + input2[f'c{i}'].value*tmp)
        if sum1 >= 16 and sum2 == 9.5 and (input2['d16'].value + input2['e16'].value + input2['f16'].value + input2['g16'].value + input2['c16'].value*tmp) == 1:
            return 1
        else:
            return 0

    def labo_math(self, input2, tmp):
        sum_ = 0
        for i in range(2, 166):
            sum_ = sum_ + data[f'l{i}'].value * data[f'c{i}'].value * (input2[f'd{i}'].value + input2[f'e{i}'].value + input2[f'f{i}'].value + input2[f'g{i}'].value + input2[f'c{i}'].value*tmp)
        if sum_ >= 26:
            return 1
        else:
            return 0

    def labo_tech(self, input2, tmp):
        sum1 = 0
        for i in range(2, 166):
            sum1 = sum1 + data[f'o{i}'].value * data[f'c{i}'].value * (input2[f'd{i}'].value + input2[f'e{i}'].value + input2[f'f{i}'].value + input2[f'g{i}'].value + input2[f'c{i}'].value*tmp)
        sum2 = 0
        for i in range(2, 166):
            sum2 = sum2 + data[f'p{i}'].value * data[f'c{i}'].value * (input2[f'd{i}'].value + input2[f'e{i}'].value + input2[f'f{i}'].value + input2[f'g{i}'].value + input2[f'c{i}'].value*tmp)
        if sum1 == 6 and sum2 >= 20:
            return 1
        else:
            return 0

    #物理コース卒業研究
    def labo_phys(self, input2, tmp):
        sum1 = 0
        for i in range(2, 166):
            sum1 = sum1 + data[f'm{i}'].value * data[f'c{i}'].value * (input2[f'd{i}'].value + input2[f'e{i}'].value + input2[f'f{i}'].value + input2[f'g{i}'].value + input2[f'c{i}'].value*tmp)
        sum2 = 0
        for i in range(2, 166):
            sum2 = sum2 + data[f'n{i}'].value * data[f'c{i}'].value * (input2[f'd{i}'].value + input2[f'e{i}'].value + input2[f'f{i}'].value + input2[f'g{i}'].value + input2[f'c{i}'].value*tmp)
        if sum1 == 4 and sum2 >= 22:
            return 1
        else:
            return 0

    #研究室に入れるかどうか
    def labo(self, input1, input2, tmp):
        c = input1['b1'].value
        if c == 0:
            print('コース未定なので算出できません')
        elif c == 1:
            if calc.foundation(input2, tmp) == 1 and calc.labo_math(input2 ,tmp) == 1:
                print('数学コースの卒業研究が履修できます')
                return 1
        elif c == 2:
            if calc.foundation(input2, tmp) == 1 and calc.labo_tech(input2, tmp) == 1:
                print('情報コースの卒業研究が履修できます')
                return 1
        elif c == 3:
            if calc.foundation(input2, tmp) == 1 and calc.labo_math(input2, tmp) == 1:
                print('物理コースの卒業研究が履修できます')
                return 1
        print('卒業研究は履修できません')
        return 0

    #以下は卒業判定。tmp=0...履修予定を含まない / tmp=1...履修予定を含む
    #全コース共通科目
    def graduate_all(self, input4, tmp):
        #教養特別講義
        sum1 = 0
        for i in range(2,5):
            sum1 = sum1 + input4[f'b{i}'].value * (input4[f'i{i}'].value + input4[f'c{i}'].value*tmp)
        if sum1 != 2:
            return 0
        #外国語
        sum1 = 0
        for i in range(5,11):
            sum1 = sum1 + input4[f'b{i}'].value * (input4[f'd{i}'].value + input4[f'e{i}'].value + input4[f'f{i}'].value + input4[f'g{i}'].value + input4[f'c{i}'].value*tmp)
        if sum1 != 12:
            return 0
        #情報処理
        sum1 = input4['b11'].value * (input4['d11'].value + input4['e11'].value + input4['f11'].value + input4['g11'].value + input4['c11'].value*tmp)
        if sum1 != 2:
            return 0
        #身体運動
        for i in range(12,15):
            sum1 = sum1 + input4[f'b{i}'].value * (input4[f'd{i}'].value + input4[f'e{i}'].value + input4[f'f{i}'].value + input4[f'g{i}'].value + input4[f'c{i}'].value*tmp)
        if sum1 != 2:
            return 0
        #教養科目/その他科目
        sum1 = 0 #教養科目の単位数
        sum2 = 0 #その他科目の単位数
        i = 15
        cnt_a = 0 #教養Aのカウント
        cnt_b = 0 #教養Bのカウント
        cnt_c = 0 #教養Cのカウント
        while input4[f'a{i}'].value != None:
            if input4[f'l{i}'].value == 1:
                cnt_a = cnt_a + 1 
                sum1 = sum1 + input4[f'b{i}'].value * (input4[f'd{i}'].value + input4[f'e{i}'].value + input4[f'f{i}'].value + input4[f'g{i}'].value + input4[f'c{i}'].value*tmp)                
            elif input4[f'm{i}'].value == 1:
                cnt_b = cnt_b + 1
                sum1 = sum1 + input4[f'b{i}'].value * (input4[f'd{i}'].value + input4[f'e{i}'].value + input4[f'f{i}'].value + input4[f'g{i}'].value + input4[f'c{i}'].value*tmp)                
            elif input4[f'n{i}'].value == 1:
                cnt_c = cnt_c + 1
                sum1 = sum1 + input4[f'b{i}'].value * (input4[f'd{i}'].value + input4[f'e{i}'].value + input4[f'f{i}'].value + input4[f'g{i}'].value + input4[f'c{i}'].value*tmp)                
            else:
                sum2 = sum2 + input4[f'b{i}'].value * (input4[f'd{i}'].value + input4[f'e{i}'].value + input4[f'f{i}'].value + input4[f'g{i}'].value + input4[f'i{i}'].value + input4[f'c{i}'].value*tmp)                
            i = i + 1
        if cnt_a == 0 or cnt_b == 0 or cnt_c == 0 or sum1 < 12:
            return 0
        else:
            return sum1 + sum2 - 12

    #数学コース卒業判定
    def graduate_math(self, input2, tmp):
        sum1 = 0
        for i in range(2, 166):
            sum1 = sum1 + data[f'i{i}'].value * data[f'c{i}'].value * (input2[f'd{i}'].value + input2[f'e{i}'].value + input2[f'f{i}'].value + input2[f'g{i}'].value + input2[f'c{i}'].value*tmp)
        sum2 = 0
        for i in range(23, 166):
            sum2 = sum2 + data[f'c{i}'].value * (input2[f'd{i}'].value + input2[f'e{i}'].value + input2[f'f{i}'].value + input2[f'g{i}'].value + input2[f'c{i}'].value*tmp)
        if calc.foundation(input2, tmp) == 1 and calc.labo_math(input2 ,tmp) == 1 and sum1 >= 40 and sum2 - 40 >= 17.5:
            return sum2 - (40 + 17.5)
        else:
            return 0

    #情報コース卒業判定
    def graduate_tech(self, input2, tmp):
        sum1 = 0
        for i in range(2, 166):
            sum1 = sum1 + data[f'w{i}'].value * data[f'c{i}'].value * (input2[f'd{i}'].value + input2[f'e{i}'].value + input2[f'f{i}'].value + input2[f'g{i}'].value + input2[f'c{i}'].value*tmp)
        sum2 = 0
        for i in range(2, 166):
            sum2 = sum2 + data[f'h{i}'].value * data[f'c{i}'].value * (input2[f'd{i}'].value + input2[f'e{i}'].value + input2[f'f{i}'].value + input2[f'g{i}'].value + input2[f'c{i}'].value*tmp)
        sum3 = 0
        for i in range(23, 166):
            sum3 = sum3 + data[f'c{i}'].value * (input2[f'd{i}'].value + input2[f'e{i}'].value + input2[f'f{i}'].value + input2[f'g{i}'].value + input2[f'c{i}'].value*tmp)
        if sum1 >= 13 and sum2 - 13 >= 30 and sum3 - (13 + 30) >= 14.5 and calc.foundation(input2, tmp) == 1 and calc.labo_tech(input2, tmp) == 1:
            return sum3 - (13 + 30 + 14.5)
        else:
            return 0

    #物理コース卒業判定
    def graduate_phys(self, input2, tmp):
        sum1 = 0
        for i in range(2, 166):
            sum1 = sum1 + data[f'v{i}'].value * data[f'c{i}'].value * (input2[f'd{i}'].value + input2[f'e{i}'].value + input2[f'f{i}'].value + input2[f'g{i}'].value + input2[f'c{i}'].value*tmp)
        sum2 = 0
        for i in range(2, 166):
            sum2 = sum2 + data[f'g{i}'].value * data[f'c{i}'].value * (input2[f'd{i}'].value + input2[f'e{i}'].value + input2[f'f{i}'].value + input2[f'g{i}'].value + input2[f'c{i}'].value*tmp)
        sum3 = 0
        for i in range(23, 166):
            sum3 = sum3 + data[f'c{i}'].value * (input2[f'd{i}'].value + input2[f'e{i}'].value + input2[f'f{i}'].value + input2[f'g{i}'].value + input2[f'c{i}'].value*tmp)
        if sum1 >= 23.5 and sum2 - 23.5 >= 20 and sum3 - (23.5 + 20) >= 14 and calc.foundation(input2, tmp) == 1 and calc.labo_phys(input2, tmp) == 1:
            return sum3 - (23.5 + 20 + 14)
        else:
            return 0

    #卒業できるか
    def graduate(self, input1, input2, input4, tmp):
        c = input1['b1'].value
        if c == 0:
            print('コース未定なので算出できません')
        elif c == 1:
            if calc.graduate_all(input4, tmp) + calc.graduate_math(input2, tmp) >= 5:
                print('卒業できます(数学コース)')
                return 1
        elif c == 2:
            if calc.graduate_all(input4, tmp) + calc.graduate_tech(input2, tmp) >= 5:
                print('卒業できます(情報コース)')
                return 1
        elif c == 3:
            if calc.graduate_all(input4, tmp) + calc.graduate_phys(input2, tmp) >= 5:
                print('卒業できます(物理コース)')
                return 1
        print('卒業できません')
        return 0

    #教職科目の判定
    #modeの対応は1~5の順で数学中、数学高、情報、理科中、理科高
    #法律で定められた科目
    def tp_1(self, input4, tmp):
        #プレゼンテーションイングリッシュb
        sum_ = 0
        for i in {6, 9}:
            sum_ = sum_ + input4[f'b{i}'].value * input4[f'b{i}'].value * (input4[f'd{i}'].value + input4[f'e{i}'].value + input4[f'f{i}'].value + input4[f'g{i}'].value + input4[f'i{i}'].value + input4[f'c{i}'].value*tmp)                
        if sum_ < 2:
            return 0
        #身体運動
        sum_ = 0
        for i in range(12, 15):
            sum_ = sum_ + input4[f'b{i}'].value * input4[f'b{i}'].value * (input4[f'd{i}'].value + input4[f'e{i}'].value + input4[f'f{i}'].value + input4[f'g{i}'].value + input4[f'i{i}'].value + input4[f'c{i}'].value*tmp)                
        if sum_ < 2:
            return 0            
        #基礎情報処理
        i = 11
        if input4[f'b{i}'].value * input4[f'b{i}'].value * (input4[f'd{i}'].value + input4[f'e{i}'].value + input4[f'f{i}'].value + input4[f'g{i}'].value + input4[f'i{i}'].value + input4[f'c{i}'].value*tmp) != 2:
            return 0    
        #日本国憲法
        i = 15
        if input4[f'b{i}'].value * input4[f'b{i}'].value * (input4[f'd{i}'].value + input4[f'e{i}'].value + input4[f'f{i}'].value + input4[f'g{i}'].value + input4[f'i{i}'].value + input4[f'c{i}'].value*tmp) != 2:
            return 0    
        return 1

    #教職に関する科目
    def tp_2(self, input3, tmp, mode):
        mode_dict = {
            1: 'd',
            2: 'e',
            3: 'f',
            4: 'g',
            5: 'h'
        }
        sum_ = 0
        for i in range(2, 23):
            sum_ = sum_ + data_t[f'c{i}'].value * data_t[f'{mode_dict[mode]}{i}'] * (input3[f'd{i}'].value + input3[f'e{i}'].value + input3[f'f{i}'].value + input3[f'g{i}'].value + input3[f'c{i}'].value*tmp)
        if mode in {1, 4}:
            if sum_ < 31:
                return 0
        elif mode in {2, 3, 5}:
            if sum_ < 27:
                return 0
        return 1

    #教科に関する科目:
    def tp_3(self, input2, tmp, mode):
        mode_dict1_1 = {
            1: 'x',
            2: 'x',
            3: 'z',
            4: 'ab',
            5: 'ab'
        }
        mode_dict1_2 = {
            1: 'y',
            2: 'y',
            3: 'aa',
            4: 'ac',
            5: 'ac'
        }
        mode_dict2 = {
            1: 28,
            2: 32,
            3: 34,
            4: 28,
            5: 32
        }
        sum_ = 0
        for i in range(2, 166):
            sum_ = sum_ + data[f'c{i}'].value * data[f'{mode_dict1_1[mode]}{i}'] * data[f'{mode_dict1_2[mode]}{i}'] * (input2[f'd{i}'].value + input2[f'e{i}'].value + input2[f'f{i}'].value + input2[f'g{i}'].value + input2[f'c{i}'].value*tmp)
        if sum_ < mode_dict2[mode]:
            return 0
        return 1

    def tp_calc(self, input2, input3, input4, tmp):
        mode_dict = {
            1: '数学中',
            2: '数学高',
            3: '情報',
            4: '理科中',
            5: '理科高'
        }
        cnt = 0
        for mode in range(1, 6):
            if calc.tp_1(input4, tmp) == 1 and calc.tp_2(input3, tmp, mode) == 1 and calc.tp_3(input2, tmp, mode) == 1:
                print(f'{mode_dict[mode]}の教職が取得できる予定です')
                cnt = cnt + 1
        if cnt == 0:
            print('取得できる予定の教職はありません')
calc = calc()

def main():
    while(True):
        func.txt('start')
        c1 = int(input())
        if c1 in {1, 2, 3, 4}:
            #1: アカウント新規作成
            if c1 == 1:
                while(True):
                    if user.new() == 1:
                        break
            #2: ログイン
            elif c1 == 2:
                logindata = user.login()
                if logindata != 0:
                    #ユーザーデータの読み込み
                    input_ = openpyxl.load_workbook(f'./data/{logindata[1]}/input.xlsx')
                    input1 = input_['input1']
                    input2 = input_['input2']
                    input3 = input_['input3']
                    input4 = input_['input4']
                    def inputsave():    #保存用関数（引数として使う）
                        input_.save(f'./data/{logindata[1]}/input.xlsx')
                    print(f'ようこそ {logindata[1]} さん')
                    while(True):
                        func.txt('after_login')
                        c2 = int(input())
                        if c2 in {1, 2, 3}:
                            #1: ユーザー情報・成績確認
                            if c2 == 1:
                                print('ユーザーネーム:', f'{logindata[1]}')
                                print('コース:', user.corse_show(input1))
                                print('学年:', user.grade_show(input1))
                                print('取得希望の教職')
                                user.tp_show(input1)
                                print('学科科目GPA:', calc.gpa(input2))
                                print('GPA:', calc.gpaall(input2, input3, input4))
                                print('卒業研究(履修予定含む)')
                                calc.labo(input1, input2, 1)
                                print('卒業研究(履修予定含まない)')
                                calc.labo(input1, input2, 0)
                                print('卒業判定(履修予定含む)')
                                calc.graduate(input1, input2, input4, 1)
                                print('卒業判定(履修予定含まない)')
                                calc.graduate(input1, input2, input4, 0)
                                print('教職判定(履修予定含む)')
                                calc.tp_calc(input2, input3, input4, 1)
                                print('教職判定(履修予定含まない)')
                                calc.tp_calc(input2, input3, input4, 0)
                            #2: アカウント情報変更
                            elif c2 == 2:
                                while(True):
                                    func.txt('account_info')
                                    c3 = int(input())
                                    if c3 in {1, 2, 3, 4}:
                                        #1: コース
                                        if c3 == 1:
                                            user.corse(input1, inputsave)
                                        #2: 学年
                                        elif c3 == 2:
                                            user.grade(input1, inputsave)
                                        #3: 教職
                                        elif c3 == 3:
                                            user.tp(input1, inputsave)
                                        elif c3 == 4:
                                            break
                                    else:
                                        print('ERROR: もう一度入力し直してください')
                            #3: ログアウト
                            elif c2 == 3:
                                print('ログアウトしました。')
                                break
                        else:
                            print('ERROR: もう一度入力し直してください')
            #3: アカウント削除
            elif c1 == 3:
                while('True'):
                    if user.delete() >= 1:
                        break
            #4: 終了する
            elif c1 == 4:
                print('**終了**')
                break
        else:
            print('ERROR: もう一度入力し直してください')

if __name__ == "__main__":
    main()